"""
RCF Export — Multi-format export engine.

Output formats:
  CSV  — Standard comma-separated with fixed column order
  JSON — Nested structure matching Pydantic ContactModel
  XLSX — Styled spreadsheet with conditional coloring, autofilter, frozen headers
  HTML — Jinja2 template with responsive table and confidence color-coding
"""

from __future__ import annotations

import csv
import io
import json
import os
import shutil
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from pydantic import BaseModel, Field

from models.enums import ExportFormat


# ---------------------------------------------------------------------------
# Pydantic models — the canonical data shape
# ---------------------------------------------------------------------------

class ContactModel(BaseModel):
    """
    Canonical contact record.
    This is the single source of truth for all export formats.
    """

    id: str = Field(description="Unique contact identifier (UUID).")
    first_name: str = Field(default="", description="First name.")
    last_name: str = Field(default="", description="Last name / family name.")
    full_name: str = Field(description="Full display name.")

    # Email
    email: str = Field(default="", description="Primary email address.")
    email_verified: Optional[bool] = Field(default=None, description="True if email passed verification.")
    email_verification_tier: str = Field(default="", description="Highest verification tier reached (syntax/dns/smtp/api).")
    email_sources: List[str] = Field(default_factory=list, description="Sources that provided this email.")

    # Phone
    phone: str = Field(default="", description="Primary phone number in E.164 format.")
    phone_formatted: str = Field(default="", description="Human-readable phone format.")
    phone_carrier: str = Field(default="", description="Detected carrier (Etisalat, du, Virgin Mobile, etc.).")
    phone_type: str = Field(default="", description="Mobile / Landline / VoIP.")
    phone_verified: Optional[bool] = Field(default=None, description="True if phone passed libphonenumber validation.")

    # WhatsApp
    whatsapp: Optional[bool] = Field(default=None, description="True if number is WhatsApp-enabled.")
    whatsapp_business: Optional[bool] = Field(default=None, description="True if WhatsApp Business account detected.")

    # Professional
    title: str = Field(default="", description="Job title (e.g. 'Senior Technical Recruiter').")
    company: str = Field(default="", description="Company / agency name.")
    company_domain: str = Field(default="", description="Company website domain.")
    industry: str = Field(default="", description="Industry sector.")
    linkedin_url: str = Field(default="", description="LinkedIn profile URL.")

    # Geography
    region: str = Field(default="", description="Region code (uae, saudi, etc.).")
    emirate: str = Field(default="", description="Detected emirate (for UAE contacts).")

    # Meta
    sources: List[str] = Field(default_factory=list, description="All sources that contributed data.")
    confidence_score: float = Field(default=0.0, description="Composite confidence score (0.0–1.0).")
    confidence_tier: str = Field(default="", description="high / medium / low / unverified.")
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc), description="First discovery timestamp.")
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc), description="Last enrichment/update timestamp.")


# ---------------------------------------------------------------------------
# CSV columns — fixed order
# ---------------------------------------------------------------------------

CSV_COLUMNS: List[str] = [
    "Name",
    "Email",
    "Email_Verified",
    "Email_Verification_Tier",
    "Phone",
    "Phone_Formatted",
    "Phone_Carrier",
    "Phone_Type",
    "Phone_Verified",
    "WhatsApp",
    "WhatsApp_Business",
    "Title",
    "Company",
    "Company_Domain",
    "Industry",
    "LinkedIn_URL",
    "Region",
    "Emirate",
    "Sources",
    "Confidence_Score",
    "Confidence_Tier",
    "Created_At",
    "Updated_At",
]


def contact_to_csv_row(c: ContactModel) -> Dict[str, str]:
    """Convert a ContactModel to a flat dict with CSV_COLUMNS as keys."""
    return {
        "Name": c.full_name,
        "Email": c.email,
        "Email_Verified": _bool_str(c.email_verified),
        "Email_Verification_Tier": c.email_verification_tier,
        "Phone": c.phone,
        "Phone_Formatted": c.phone_formatted,
        "Phone_Carrier": c.phone_carrier,
        "Phone_Type": c.phone_type,
        "Phone_Verified": _bool_str(c.phone_verified),
        "WhatsApp": _bool_str(c.whatsapp),
        "WhatsApp_Business": _bool_str(c.whatsapp_business),
        "Title": c.title,
        "Company": c.company,
        "Company_Domain": c.company_domain,
        "Industry": c.industry,
        "LinkedIn_URL": c.linkedin_url,
        "Region": c.region,
        "Emirate": c.emirate,
        "Sources": "; ".join(c.sources),
        "Confidence_Score": f"{c.confidence_score:.2f}",
        "Confidence_Tier": c.confidence_tier,
        "Created_At": c.created_at.isoformat(),
        "Updated_At": c.updated_at.isoformat(),
    }


def _bool_str(val: Optional[bool]) -> str:
    if val is None:
        return ""
    return "TRUE" if val else "FALSE"


# ---------------------------------------------------------------------------
# Export functions
# ---------------------------------------------------------------------------

def export_csv(
    contacts: Sequence[ContactModel],
    output: Path,
) -> Path:
    """
    Export contacts as CSV with fixed column order.

    UTF-8 with BOM for Excel compatibility. Columns match CSV_COLUMNS exactly.
    Uses atomic write to prevent corrupt files on crash.
    """
    output = Path(output)
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".csv", dir=output.parent)
    try:
        with open(tmp_fd, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=CSV_COLUMNS)
            writer.writeheader()
            for contact in contacts:
                writer.writerow(contact_to_csv_row(contact))
        shutil.move(tmp_path, output)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise
    return output


def export_json(
    contacts: Sequence[ContactModel],
    output: Path,
) -> Path:
    """
    Export contacts as JSON matching Pydantic model structure.

    Uses atomic write to prevent corrupt files on crash.

    Output shape:
    {
      "generated_at": "2026-05-04T12:00:00",
      "total_contacts": 42,
      "contacts": [ { ... ContactModel ... }, ... ]
    }
    """
    output = Path(output)
    payload = {
        "generated_at": datetime.now().isoformat(),
        "total_contacts": len(contacts),
        "contacts": [c.model_dump(mode="json") for c in contacts],
    }
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".json", dir=output.parent)
    try:
        with open(tmp_fd, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2, ensure_ascii=False, default=str)
        shutil.move(tmp_path, output)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise
    return output


def export_xlsx(
    contacts: Sequence[ContactModel],
    output: Path,
) -> Path:
    """
    Export contacts as a styled XLSX spreadsheet.

    Features:
    - Header row: bold white text on dark blue, frozen
    - Autofilter on all columns
    - Auto-width columns
    - Conditional coloring by confidence tier:
        high (>=0.8) = green, medium (0.5-0.79) = yellow, low (<0.5) = red
    - Confidence_Score column formatted as percentage

    Requires: openpyxl
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for XLSX export. Install with: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Recruiter Contacts"

    # ── Header styling ─────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write headers
    for col_idx, header in enumerate(CSV_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Freeze top row
    ws.freeze_panes = "A2"

    # ── Confidence tier fills ──────────────────────────────────────────
    tier_fills = {
        "high": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # green
        "medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # yellow
        "low": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # red
        "unverified": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),  # grey
    }

    # ── Data rows ──────────────────────────────────────────────────────
    for row_idx, contact in enumerate(contacts, start=2):
        row_data = contact_to_csv_row(contact)
        for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

            # Apply tier coloring to the entire row based on confidence tier
            tier = contact.confidence_tier.lower() if contact.confidence_tier else "unverified"
            if tier in tier_fills:
                cell.fill = tier_fills[tier]

        # Format confidence score as percentage
        score_col = CSV_COLUMNS.index("Confidence_Score") + 1
        score_cell = ws.cell(row=row_idx, column=score_col)
        try:
            score_cell.value = float(score_cell.value)
            score_cell.number_format = "0.00%"
        except (ValueError, TypeError):
            pass

    # ── Auto-width columns ─────────────────────────────────────────────
    for col_idx in range(1, len(CSV_COLUMNS) + 1):
        max_length = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(len(contacts) + 2, 100))  # sample first 100 rows
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    # ── Autofilter ─────────────────────────────────────────────────────
    last_col = get_column_letter(len(CSV_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(contacts) + 1}"

    # Atomic write: save to temp file first, then rename
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=output.parent)
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
        shutil.move(tmp_path, output)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise
    return output


def export_html(
    contacts: Sequence[ContactModel],
    output: Path,
) -> Path:
    """
    Export contacts as a responsive HTML page using Jinja2.

    Features:
    - Responsive table with horizontal scroll on mobile
    - Confidence tier color-coding (green/yellow/red)
    - Clickable email and LinkedIn links
    - Sortable columns (via JavaScript)
    - Print-friendly styles
    """
    try:
        from jinja2 import Environment, FileSystemLoader, select_autoescape
    except ImportError:
        raise ImportError("jinja2 is required for HTML export. Install with: pip install jinja2")

    # Read the template (inline fallback if template file not found)
    template_str = HTML_TEMPLATE

    env = Environment(autoescape=select_autoescape(["html"]))
    template = env.from_string(template_str)

    rows = [contact_to_csv_row(c) for c in contacts]
    html_content = template.render(
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        total_contacts=len(contacts),
        columns=CSV_COLUMNS,
        contacts=rows,
    )

    # Atomic write
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".html", dir=output.parent)
    try:
        with open(tmp_fd, "w", encoding="utf-8") as fh:
            fh.write(html_content)
        shutil.move(tmp_path, output)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise
    return output


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

EXPORTERS = {
    ExportFormat.CSV: export_csv,
    ExportFormat.JSON: export_json,
    ExportFormat.XLSX: export_xlsx,
    ExportFormat.HTML: export_html,
}


def export_contacts(
    contacts: Sequence[ContactModel],
    fmt: ExportFormat,
    output: Optional[Path] = None,
) -> Path:
    """
    Export contacts in the specified format.

    If output is None, auto-generates a filename:
        rcf_export_20260504_120000.{ext}

    Returns the path to the exported file.
    """
    if output is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ext = fmt.value
        output = Path(f"rcf_export_{timestamp}.{ext}")

    exporter = EXPORTERS.get(fmt)
    if exporter is None:
        raise ValueError(f"Unsupported export format: {fmt}")

    return exporter(contacts, output)


# ---------------------------------------------------------------------------
# HTML Jinja2 template (inline — no external file dependency)
# ---------------------------------------------------------------------------

HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>RCF — Recruiter Contacts Export</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f7fa; color: #333; padding: 20px; }
  h1 { font-size: 1.5rem; margin-bottom: 0.25rem; color: #1F4E79; }
  .meta { font-size: 0.85rem; color: #666; margin-bottom: 1rem; }
  .table-wrap { overflow-x: auto; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }
  table { width: 100%; border-collapse: collapse; background: #fff; font-size: 0.85rem; }
  th { background: #1F4E79; color: #fff; padding: 10px 12px; text-align: left; white-space: nowrap; cursor: pointer; position: sticky; top: 0; }
  th:hover { background: #163A5C; }
  td { padding: 8px 12px; border-bottom: 1px solid #e8e8e8; white-space: nowrap; }
  tr:hover td { background: #f0f4f8; }
  .tier-high { background: #C6EFCE !important; }
  .tier-medium { background: #FFEB9C !important; }
  .tier-low { background: #FFC7CE !important; }
  .tier-unverified { background: #D9D9D9 !important; }
  a { color: #1F4E79; text-decoration: none; }
  a:hover { text-decoration: underline; }
  @media print {
    body { padding: 0; background: #fff; }
    .table-wrap { box-shadow: none; }
    th { background: #333 !important; -webkit-print-color-adjust: exact; print-color-adjust: exact; }
    .tier-high, .tier-medium, .tier-low, .tier-unverified { -webkit-print-color-adjust: exact; print-color-adjust: exact; }
  }
</style>
</head>
<body>
<h1>Recruiter Contacts Export</h1>
<p class="meta">Generated: {{ generated_at }} | Total: {{ total_contacts }} contacts</p>
<div class="table-wrap">
<table>
<thead>
<tr>
  {% for col in columns %}
  <th>{{ col }}</th>
  {% endfor %}
</tr>
</thead>
<tbody>
{% for contact in contacts %}
{% set tier = contact.get('Confidence_Tier', 'unverified').lower() %}
<tr class="tier-{{ tier }}">
  {% for col in columns %}
  <td>
    {% if col == 'Email' and contact.get(col) %}
      <a href="mailto:{{ contact[col] }}">{{ contact[col] }}</a>
    {% elif col == 'LinkedIn_URL' and contact.get(col) %}
      <a href="{{ contact[col] }}" target="_blank">{{ contact[col] }}</a>
    {% elif col == 'WhatsApp' and contact[col] == 'TRUE' %}
      📱 ✅
    {% elif col == 'WhatsApp_Business' and contact[col] == 'TRUE' %}
      🏢 ✅
    {% else %}
      {{ contact.get(col, '') }}
    {% endif %}
  </td>
  {% endfor %}
</tr>
{% endfor %}
</tbody>
</table>
</div>
<script>
// Simple client-side sort
document.querySelectorAll('th').forEach((th, colIdx) => {
  th.addEventListener('click', () => {
    const tbody = document.querySelector('tbody');
    const rows = Array.from(tbody.querySelectorAll('tr'));
    const asc = th.dataset.sort !== 'asc';
    document.querySelectorAll('th').forEach(h => delete h.dataset.sort);
    th.dataset.sort = asc ? 'asc' : 'desc';
    rows.sort((a, b) => {
      const aVal = a.cells[colIdx].textContent.trim();
      const bVal = b.cells[colIdx].textContent.trim();
      return asc ? aVal.localeCompare(bVal) : bVal.localeCompare(aVal);
    });
    rows.forEach(r => tbody.appendChild(r));
  });
});
</script>
</body>
</html>
"""
