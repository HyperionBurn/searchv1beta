"""
Tests for the export engine (rcf/export.py).

Covers:
  - CSV export with sample contacts
  - JSON export with sample contacts
  - Verify output file contents contain expected data
"""

import json
import tempfile
from datetime import datetime
from pathlib import Path

import pytest

from rcf.export import (
    ContactModel,
    CSV_COLUMNS,
    contact_to_csv_row,
    export_csv,
    export_json,
    export_xlsx,
)


# ===========================================================================
# Test data
# ===========================================================================


def _make_contact(**overrides) -> ContactModel:
    """Create a sample ContactModel with sensible defaults."""
    defaults = {
        "id": "test-uuid-001",
        "full_name": "Ahmed Al-Rashid",
        "first_name": "Ahmed",
        "last_name": "Al-Rashid",
        "email": "ahmed@dmcc.ae",
        "email_verified": True,
        "email_verification_tier": "smtp",
        "phone": "+971501234567",
        "phone_formatted": "+971 50 123 4567",
        "phone_carrier": "etisalat",
        "phone_type": "Mobile",
        "phone_verified": True,
        "whatsapp": True,
        "whatsapp_business": False,
        "title": "Senior Recruiter",
        "company": "DMCC",
        "company_domain": "dmcc.ae",
        "industry": "Recruitment",
        "linkedin_url": "https://linkedin.com/in/ahmed-rashid",
        "region": "uae",
        "emirate": "Dubai",
        "sources": ["linkedin", "google_maps"],
        "confidence_score": 0.85,
        "confidence_tier": "high",
        "created_at": datetime(2026, 4, 15, 10, 0, 0),
        "updated_at": datetime(2026, 4, 15, 10, 0, 0),
    }
    defaults.update(overrides)
    return ContactModel(**defaults)


SAMPLE_CONTACTS = [
    _make_contact(),
    _make_contact(
        id="test-uuid-002",
        full_name="Saeed Khan",
        first_name="Saeed",
        last_name="Khan",
        email="saeed@example.com",
        email_verified=False,
        email_verification_tier="",
        phone="+971559876543",
        phone_carrier="du",
        confidence_score=0.55,
        confidence_tier="medium",
        sources=["bayt"],
    ),
]


# ===========================================================================
# ContactModel
# ===========================================================================


class TestContactModel:
    def test_creation(self):
        c = _make_contact()
        assert c.full_name == "Ahmed Al-Rashid"
        assert c.email == "ahmed@dmcc.ae"
        assert c.confidence_score == 0.85

    def test_model_dump(self):
        c = _make_contact()
        data = c.model_dump()
        assert data["full_name"] == "Ahmed Al-Rashid"
        assert "id" in data


# ===========================================================================
# contact_to_csv_row
# ===========================================================================


class TestContactToCSVRow:
    def test_all_columns_present(self):
        c = _make_contact()
        row = contact_to_csv_row(c)
        for col in CSV_COLUMNS:
            assert col in row, f"Missing column: {col}"

    def test_name_in_row(self):
        c = _make_contact()
        row = contact_to_csv_row(c)
        assert row["Name"] == "Ahmed Al-Rashid"

    def test_email_in_row(self):
        c = _make_contact()
        row = contact_to_csv_row(c)
        assert row["Email"] == "ahmed@dmcc.ae"

    def test_bool_values(self):
        c = _make_contact(email_verified=True)
        row = contact_to_csv_row(c)
        assert row["Email_Verified"] == "TRUE"

        c2 = _make_contact(email_verified=False)
        row2 = contact_to_csv_row(c2)
        assert row2["Email_Verified"] == "FALSE"

    def test_none_bool_values(self):
        c = _make_contact(whatsapp=None)
        row = contact_to_csv_row(c)
        assert row["WhatsApp"] == ""

    def test_confidence_score_formatted(self):
        c = _make_contact(confidence_score=0.85)
        row = contact_to_csv_row(c)
        assert row["Confidence_Score"] == "0.85"

    def test_sources_joined(self):
        c = _make_contact(sources=["linkedin", "google_maps"])
        row = contact_to_csv_row(c)
        assert "linkedin" in row["Sources"]
        assert "google_maps" in row["Sources"]


# ===========================================================================
# export_csv
# ===========================================================================


class TestExportCSV:
    def test_creates_file(self, tmp_path):
        output = tmp_path / "test.csv"
        export_csv(SAMPLE_CONTACTS, output)
        assert output.exists()

    def test_file_not_empty(self, tmp_path):
        output = tmp_path / "test.csv"
        export_csv(SAMPLE_CONTACTS, output)
        content = output.read_text(encoding="utf-8-sig")
        assert len(content) > 0

    def test_contains_header(self, tmp_path):
        output = tmp_path / "test.csv"
        export_csv(SAMPLE_CONTACTS, output)
        content = output.read_text(encoding="utf-8-sig")
        first_line = content.split("\n")[0]
        assert "Name" in first_line
        assert "Email" in first_line
        assert "Phone" in first_line

    def test_contains_contact_data(self, tmp_path):
        output = tmp_path / "test.csv"
        export_csv(SAMPLE_CONTACTS, output)
        content = output.read_text(encoding="utf-8-sig")
        assert "Ahmed Al-Rashid" in content
        assert "ahmed@dmcc.ae" in content
        assert "Saeed Khan" in content

    def test_empty_contacts_creates_header_only(self, tmp_path):
        output = tmp_path / "empty.csv"
        export_csv([], output)
        content = output.read_text(encoding="utf-8-sig")
        lines = [l for l in content.strip().split("\n") if l.strip()]
        assert len(lines) == 1  # only header

    def test_utf8_bom_present(self, tmp_path):
        output = tmp_path / "bom.csv"
        export_csv(SAMPLE_CONTACTS, output)
        raw = output.read_bytes()
        assert raw[:3] == b"\xef\xbb\xbf"  # UTF-8 BOM


# ===========================================================================
# export_json
# ===========================================================================


class TestExportJSON:
    def test_creates_file(self, tmp_path):
        output = tmp_path / "test.json"
        export_json(SAMPLE_CONTACTS, output)
        assert output.exists()

    def test_valid_json(self, tmp_path):
        output = tmp_path / "test.json"
        export_json(SAMPLE_CONTACTS, output)
        data = json.loads(output.read_text(encoding="utf-8"))
        assert "contacts" in data
        assert "generated_at" in data
        assert "total_contacts" in data

    def test_total_contacts_count(self, tmp_path):
        output = tmp_path / "test.json"
        export_json(SAMPLE_CONTACTS, output)
        data = json.loads(output.read_text(encoding="utf-8"))
        assert data["total_contacts"] == 2

    def test_contact_data_present(self, tmp_path):
        output = tmp_path / "test.json"
        export_json(SAMPLE_CONTACTS, output)
        data = json.loads(output.read_text(encoding="utf-8"))
        names = [c["full_name"] for c in data["contacts"]]
        assert "Ahmed Al-Rashid" in names
        assert "Saeed Khan" in names

    def test_contact_fields(self, tmp_path):
        output = tmp_path / "test.json"
        export_json(SAMPLE_CONTACTS, output)
        data = json.loads(output.read_text(encoding="utf-8"))
        first = data["contacts"][0]
        assert "email" in first
        assert "phone" in first
        assert "company" in first
        assert "confidence_score" in first

    def test_empty_contacts(self, tmp_path):
        output = tmp_path / "empty.json"
        export_json([], output)
        data = json.loads(output.read_text(encoding="utf-8"))
        assert data["total_contacts"] == 0
        assert data["contacts"] == []


# ===========================================================================
# export_xlsx — confidence score formatting
# ===========================================================================


class TestExportXLSX:
    def test_xlsx_creates_file(self, tmp_path):
        """export_xlsx should create a valid .xlsx file."""
        output = tmp_path / "test.xlsx"
        export_xlsx(SAMPLE_CONTACTS, output)
        assert output.exists()
        assert output.stat().st_size > 0

    def test_xlsx_confidence_score_is_percentage(self, tmp_path):
        """Confidence score 0.85 should display as 85% not 0.85%.

        The XLSX exporter sets the cell value to the raw float (0.85)
        and applies the '0.00%' number format, so Excel renders it as 85%.
        """
        output = tmp_path / "test.xlsx"
        export_xlsx(SAMPLE_CONTACTS, output)

        from openpyxl import load_workbook

        wb = load_workbook(output)
        ws = wb.active

        # Find the Confidence_Score column index (1-based)
        score_col = None
        for col_idx in range(1, len(CSV_COLUMNS) + 1):
            if ws.cell(row=1, column=col_idx).value == "Confidence_Score":
                score_col = col_idx
                break
        assert score_col is not None, "Confidence_Score column not found in headers"

        # First data row (row 2) = Ahmed Al-Rashid with score 0.85
        score_cell = ws.cell(row=2, column=score_col)
        assert isinstance(score_cell.value, float)
        assert score_cell.value == 0.85
        # Verify the number format is percentage, not plain text
        assert "%" in score_cell.number_format

    def test_xlsx_empty_contacts(self, tmp_path):
        """export_xlsx with no contacts should create a file with just headers."""
        output = tmp_path / "empty.xlsx"
        export_xlsx([], output)
        assert output.exists()

        from openpyxl import load_workbook

        wb = load_workbook(output)
        ws = wb.active
        # Row 1 = headers, no data rows
        assert ws.max_row == 1
